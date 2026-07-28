"""
Mini Project Exhibition 2026 — Evaluation Rubric Sheets (v3)
Course : Dynamic System Modeling (702MH0C023), B.Tech Mechatronics Sem IV
Date   : 25 April 2026 (Saturday) | AR/VR Lab, MPSTME, NMIMS

6 Judges — 100 Marks Total:
  1. Mr. Kamlesh Panchal    – Industry & Manufacturing Expert      /16  (4 x 4) [MORNING]
  2. Ms. Hasti Chandarana   – Intellectual Property (IP) Expert    /16  (4 x 4) [MORNING]
  3. Dr./Ms. Nitu Gupta     – Mathematics Expert                   /16  (4 x 4) [MORNING]
  4. Mr. Mahendra Kane      – Industrial Automation / Industry 4.0 /20  (4 x 5) [MORNING]
  5. Ms. Parminder Jandoo   – Communication & Presentation Expert  /16  (4 x 4) [MORNING + AFTERNOON]
  6. Dr. Debasis Dash       – HR & Behavioural Science Expert      /16  (4 x 4) [MORNING]

Output files (both saved to DSM_Exhibition_2026 folder):
  Evaluation_Rubric_Sheets.xlsx         — digital, with SUM formulas + summary sheet
  Evaluation_Rubric_Sheets_PRINT.xlsx   — print-ready, blank total cells, no summary

Sheet structure:
  Evaluator sheets  : 7 header rows (row 7 = col headers) + group header rows + individual student rows
  DATA_START = 8    : Row 8 = Group 1 header; row 9+ = students
  Columns A-H       : Roll No | Name | C1 | C2 | C3 | C4 | Total | Comments
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import os

OUT_DIR = r"D:\Sunny\Paper\DSM_Exhibition_Framework\DSM_Exhibition_2026"
os.makedirs(OUT_DIR, exist_ok=True)

# =============================================================================
# GROUP & STUDENT DATA  (17 groups, 54 students)
# =============================================================================
# NOTE (public-repo redaction): student full names have been replaced with
# their de-identified roll number (e.g., "Student H079") for this public
# repository, consistent with the anonymization already applied to survey
# and analysis data in the companion repository. The original workbook used
# internally by the course team retains real names for administrative
# purposes only; that version is not published.
GROUPS = [
    {
        "num": 1,
        "title": "Mathematical Modeling and Dynamic Analysis of a Multi-Terrain Four-Wheeled Robot",
        "students": [("H079","Student H079"),("H082","Student H082"),("H084","Student H084")],
    },
    {
        "num": 2,
        "title": "Modeling Rocket Landing - Docking",
        "students": [("H004","Student H004"),("H005","Student H005"),("H028","Student H028"),("H031","Student H031")],
    },
    {
        "num": 3,
        "title": "Dynamic Modeling and Stability Analysis of 2-DOF Leg Mechanism for Quadruped Robot",
        "students": [("H009","Student H009"),("H012","Student H012"),("H033","Student H033")],
    },
    {
        "num": 4,
        "title": "Critical Skid Velocity Mathematical Modeling",
        "students": [("H014","Student H014"),("H015","Student H015"),("H016","Student H016"),("H025","Student H025")],
    },
    {
        "num": 5,
        "title": "Humanoid Robot - Dynamic Modeling",
        "students": [("H003","Student H003"),("H006","Student H006"),("H024","Student H024"),("H029","Student H029")],
    },
    {
        "num": 6,
        "title": "Dynamic Modeling of Hydrogen Fuel Cell-Battery Hybrid Powertrain for Electric UAV",
        "students": [("H001","Student H001"),("H002","Student H002")],
    },
    {
        "num": 7,
        "title": "Hexapod Robot - Mathematical Modeling",
        "students": [("H021","Student H021"),("H022","Student H022"),("H023","Student H023")],
    },
    {
        "num": 8,
        "title": "Mathematical Modeling and Dynamics of Ball Balancing Robot using 6-DOF Stewart Mechanism",
        "students": [("H059","Student H059"),("H060","Student H060"),("H062","Student H062")],
    },
    {
        "num": 9,
        "title": "Dynamic Modelling of Sensor-Integrated Remotely Operated Underwater Vehicle (ROV)",
        "students": [("H035","Student H035"),("H040","Student H040"),("H046","Student H046"),("H047","Student H047")],
    },
    {
        "num": 10,
        "title": "Mathematical Modeling of Motors",
        "students": [("H036","Student H036"),("H050","Student H050"),("H051","Student H051"),("H052","Student H052")],
    },
    {
        "num": 11,
        "title": "Digital Twin (DT) of Motors",
        "students": [("H020","Student H020"),("H096","Student H096")],
    },
    {
        "num": 12,
        "title": "Mathematical Modeling for Robotic Gripper Mechanism",
        "students": [("H076","Student H076"),("H078","Student H078"),("H088","Student H088")],
    },
    {
        "num": 13,
        "title": "Modeling Energy Loss in a Shock Absorber",
        "students": [("H068","Student H068"),("H071","Student H071"),("H086","Student H086"),("H097","Student H097")],
    },
    {
        "num": 14,
        "title": "Self-Balancing Robot with 2-DOF",
        "students": [("H093","Student H093"),("H095","Student H095")],
    },
    {
        "num": 15,
        "title": "Automated EoT (Electric Overhead Travel) Crane",
        "students": [("H066","Student H066"),("H067","Student H067"),("H090","Student H090"),("H094","Student H094")],
    },
    {
        "num": 16,
        "title": "Mathematical Operation of Pick-and-Place Autonomous Rover",
        "students": [("H034","Student H034"),("H041","Student H041"),("H056","Student H056")],
    },
    {
        "num": 17,
        "title": "Mathematical Modeling of Loading Effects on a Conveyor",
        "students": [("H074","Student H074"),("H075","Student H075")],
    },
]

# =============================================================================
# EVALUATOR DEFINITIONS  (6 judges)
#
# max_c     : max marks per criterion
# total_max : max total marks for this judge
# col_hdr   : short header text for the summary sheet column
# =============================================================================
EVALUATORS = [
    # ── 1. Kamlesh Panchal  (Manufacturing)  /16 ─────────────────────────────
    {
        "name": "Mr. Kamlesh Panchal",
        "col_hdr": "Kamlesh\nPanchal\n/16",
        "role": "Industry & Manufacturing Expert",
        "session": "MORNING SESSION  |  09:00 AM to 01:00 PM",
        "max_c": 4, "total_max": 16,
        "tab_color": "1F5C99", "hdr": "1F5C99", "acc": "D6E4F0",
        "focus": (
            "Manufacturing viability, scalability, and real-world industry relevance "
            "of the mechatronic systems modeled by students."
        ),
        "sfpbl": (
            "SF-PBL: Assess whether the mathematical model and simulation "
            "can inform real manufacturing and engineering decisions for each system."
        ),
        "criteria": [
            ("C1", "System Buildability  [GROUP]",
             "Components, materials, and assembly approach are practically manufacturable with current technology."),
            ("C2", "Scalability & Real-World Fit  [GROUP]",
             "Design is scalable and addresses a genuine industrial or commercial engineering problem."),
            ("C3", "Engineering Accuracy  [GROUP]",
             "Dimensions, tolerances, forces, and load parameters are physically appropriate."),
            ("C4", "Individual Technical Q&A  [INDIVIDUAL]",
             "This student explains manufacturing aspects and key design decisions clearly when probed."),
        ],
    },
    # ── 2. Hasti Chandarana  (IP)  /16 ───────────────────────────────────────
    {
        "name": "Ms. Hasti Chandarana",
        "col_hdr": "Hasti\nChandarana\n/16",
        "role": "Intellectual Property (IP) Expert",
        "session": "MORNING SESSION  |  09:00 AM to 01:00 PM",
        "max_c": 4, "total_max": 16,
        "tab_color": "7B3F91", "hdr": "7B3F91", "acc": "EFE0F5",
        "focus": (
            "Novelty, IP potential, and commercial applicability of the project concept "
            "and the mathematical modeling approach."
        ),
        "sfpbl": (
            "SF-PBL: A novel first-principles derivation may itself constitute an IP element — "
            "assess novelty of both the system choice and the mathematical formulation."
        ),
        "criteria": [
            ("C1", "Conceptual Novelty  [GROUP]",
             "Problem statement or application idea is unique and non-obvious."),
            ("C2", "IP Potential  [GROUP]",
             "Patentable elements, design rights, or trade-secret potential is identifiable."),
            ("C3", "Commercial Viability  [GROUP]",
             "Realistic market, startup, or product application exists for this system."),
            ("C4", "Individual IP Awareness  [INDIVIDUAL]",
             "This student understands IP implications of their work when questioned."),
        ],
    },
    # ── 3. Nitu Gupta  (Mathematics)  /16 ────────────────────────────────────
    {
        "name": "Dr. / Ms. Nitu Gupta",
        "col_hdr": "Nitu\nGupta\n/16",
        "role": "Mathematics Expert",
        "session": "MORNING SESSION  |  09:00 AM to 01:00 PM",
        "max_c": 4, "total_max": 16,
        "tab_color": "1A7A4A", "hdr": "1A7A4A", "acc": "D5F0E2",
        "focus": (
            "Mathematical rigour, correctness of derivation, and simulation validation — "
            "the core learning artifact of SF-PBL."
        ),
        "sfpbl": (
            "SF-PBL CORE: Mathematical derivation (ODEs, Lagrangian, state-space) is the "
            "PRIMARY artifact. Python / MATLAB code is secondary. Assess derivation first."
        ),
        "criteria": [
            ("C1", "Derivation Correctness  [GROUP]",
             "Newton / Lagrangian / state-space derivation is mathematically sound."),
            ("C2", "Equation Formulation  [GROUP]",
             "ODEs, transfer functions, and state-space matrices are correctly formulated."),
            ("C3", "Simulation Validation  [GROUP]",
             "Energy checks, parameter studies, or physical-limit verification performed."),
            ("C4", "Individual Mathematical Grasp  [INDIVIDUAL]",
             "This student explains derivation steps and answers mathematical Q&A correctly."),
        ],
    },
    # ── 4. Mahendra Kane  (Industrial Automation / Industry 4.0)  /20 ────────
    {
        "name": "Mr. Mahendra Kane",
        "col_hdr": "Mahendra\nKane\n/20",
        "role": "Industrial Automation & Industry 4.0 Expert  (Siemens Ltd)",
        "session": "MORNING SESSION  |  09:00 AM to 01:00 PM",
        "max_c": 5, "total_max": 20,
        "tab_color": "8B6914", "hdr": "8B6914", "acc": "FFF3CD",
        "focus": (
            "Industrial automation relevance, control system integration potential, and "
            "Digital Twin / Industry 4.0 readiness of the mathematical model and simulation."
        ),
        "sfpbl": (
            "SF-PBL: Assess whether the simulation-based model can bridge into real PLC / SCADA / "
            "motion-control scenarios and contribute to an Industry 4.0 digital strategy."
        ),
        "criteria": [
            ("C1", "Industrial System Relevance  [GROUP]",
             "Modeled dynamic system represents a real-world industrial automation challenge or machine system."),
            ("C2", "Control & Actuation Alignment  [GROUP]",
             "Math model is structured to connect with PLC, SCADA, servo drives, or motion control hardware."),
            ("C3", "Digital Twin / Industry 4.0 Potential  [GROUP]",
             "Simulation is formulated to serve as a Digital Twin; reflects Industry 4.0 / IoT thinking."),
            ("C4", "Individual Automation Q&A  [INDIVIDUAL]",
             "This student articulates the industrial automation relevance and control implementation strategy when questioned."),
        ],
    },
    # ── 5. Parminder Jandoo  (Communication)  /16 ────────────────────────────
    {
        "name": "Ms. Parminder Jandoo",
        "col_hdr": "Parminder\nJandoo\n/16",
        "role": "Communication & Presentation Expert",
        "session": "MORNING (09:00-13:00) + AFTERNOON (13:30-17:00)  --  Combined 16 Marks",
        "max_c": 4, "total_max": 16,
        "tab_color": "B5471B", "hdr": "B5471B", "acc": "FAE5DC",
        "focus": (
            "Communication quality, delivery, body language, and presentation skills "
            "observed across the morning demo and the afternoon formal presentation."
        ),
        "sfpbl": (
            "SF-PBL: Assess how clearly each student communicates their simulation journey: "
            "problem identification -> math model -> simulation results -> validation."
        ),
        "criteria": [
            ("C1", "Technical Clarity  [MORNING — Individual]",
             "Clear, structured explanation of the system, mathematical model, and simulation results."),
            ("C2", "Confidence & Q&A Response  [MORNING — Individual]",
             "Composure, fluency, and quality of answers during the project demo."),
            ("C3", "Presentation Delivery  [AFTERNOON — Individual]",
             "Slide quality, verbal delivery, body language, and eye contact during formal presentation."),
            ("C4", "Overall Communication Impact  [BOTH SESSIONS — Individual]",
             "Holistic communication impression observed across morning demo and afternoon presentation."),
        ],
    },
    # ── 6. Debasis Dash  (HR & Behavioural Science)  /16 ─────────────────────
    {
        "name": "Dr. Debasis Dash",
        "col_hdr": "Debasis\nDash\n/16",
        "role": "HR & Behavioural Science Expert  (MPSTME, NMIMS)",
        "session": "MORNING SESSION  |  09:00 AM to 01:00 PM",
        "max_c": 4, "total_max": 16,
        "tab_color": "1A3A6B", "hdr": "1A3A6B", "acc": "E8EAF6",
        "focus": (
            "Team collaboration, professional conduct, problem-solving mindset, and "
            "individual self-awareness within the SF-PBL group project context."
        ),
        "sfpbl": (
            "SF-PBL: Engineering education goes beyond equations — assess how well students "
            "collaborate, reflect on challenges, and demonstrate professional and ethical behaviour."
        ),
        "criteria": [
            ("C1", "Team Collaboration  [GROUP]",
             "Team demonstrates fair work distribution and effective collaboration throughout the project."),
            ("C2", "Problem-Solving Mindset  [GROUP]",
             "Team approaches challenges systematically, with persistence and reflective thinking."),
            ("C3", "Individual Professionalism  [INDIVIDUAL]",
             "This student demonstrates professional conduct, ethical awareness, and engineering responsibility."),
            ("C4", "Individual Self-Awareness  [INDIVIDUAL]",
             "This student reflects on personal contribution, shows adaptability, and communicates behavioural insights."),
        ],
    },
]

# =============================================================================
# ROW MAP
# Sheet layout identical across all evaluator sheets:
#   Rows 1–6  : header section
#   Row  7    : column-header row
#   Row  8    : Group 1 header row  ← DATA_START
#   Row  9+   : student rows (group header rows interspersed)
# =============================================================================
DATA_START = 8


def compute_row_map():
    """
    Pre-compute the Excel row number each student occupies in every evaluator
    sheet (layout is identical across all 6 sheets).
    """
    entries = []
    r = DATA_START
    for grp in GROUPS:
        r += 1                         # group-header occupies one row
        for roll, name in grp["students"]:
            entries.append({
                "grp_num": grp["num"],
                "roll": roll,
                "name": name,
                "title": grp["title"],
                "row": r,
            })
            r += 1
    return entries


STUDENT_ROW_MAP = compute_row_map()

# =============================================================================
# STYLE HELPERS
# =============================================================================

def hfill(hex_str):
    return PatternFill("solid", fgColor=hex_str)


def ap(cell, font=None, fill=None, align=None, border=None):
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border


def thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def sep():
    """Thicker bottom border to separate groups visually."""
    t = Side(style="thin",   color="BBBBBB")
    m = Side(style="medium", color="444444")
    return Border(left=t, right=t, top=t, bottom=m)


def _grading_scale_note(ev):
    """Build the grading scale + GROUP/INDIVIDUAL guidance note for a sheet footer."""
    max_c = ev["max_c"]
    if max_c == 5:
        scale = (
            "GRADING SCALE (each criterion, out of 5):   "
            "5 = Excellent  |  4 = Good  |  3 = Satisfactory  |  "
            "2 = Needs Improvement  |  1 = Poor  |  0 = Not demonstrated"
        )
    else:
        scale = (
            "GRADING SCALE (each criterion, out of 4):   "
            "4 = Excellent  |  3 = Good  |  2 = Satisfactory  |  "
            "1 = Needs Improvement  |  0 = Not demonstrated"
        )

    group_crits = [cr[0] for cr in ev["criteria"] if "group" in cr[1].lower()]
    indiv_crits = [cr[0] for cr in ev["criteria"] if "individual" in cr[1].lower()]

    if group_crits and indiv_crits:
        guidance = (
            f"  {', '.join(group_crits)}  [GROUP]: enter the SAME mark for all students in the group.   "
            f"  {', '.join(indiv_crits)}  [INDIVIDUAL]: evaluate each student SEPARATELY."
        )
    elif indiv_crits and not group_crits:
        guidance = (
            f"  All criteria are INDIVIDUAL — evaluate each student SEPARATELY based on "
            f"personal performance ({', '.join(indiv_crits)})."
        )
    else:
        guidance = (
            f"  All criteria are GROUP-level — enter the same mark for ALL students in the group "
            f"({', '.join(group_crits)})."
        )

    return scale + "\n" + guidance


# =============================================================================
# EVALUATOR SHEET BUILDER
#
# Columns (landscape A4, 8 cols A–H):
#   A  Roll No.           (10)
#   B  Student Name       (29)
#   C  C1 /<max_c>         (8)
#   D  C2 /<max_c>         (8)
#   E  C3 /<max_c>         (8)
#   F  C4 /<max_c>         (8)
#   G  Total /<total_max> (10)
#   H  Comments           (35)
# =============================================================================

def build_evaluator_sheet(ws, ev, use_formula=True):
    H     = ev["hdr"]
    A     = ev["acc"]
    max_c = ev["max_c"]
    tmax  = ev["total_max"]

    S_FILL = hfill("FFFDE7")   # score cells (light yellow)
    T_FILL = hfill("E8F5E9")   # total cell  (light green)

    # Column widths
    for col, w in {"A": 10, "B": 29, "C": 8, "D": 8,
                   "E": 8,  "F": 8,  "G": 10, "H": 35}.items():
        ws.column_dimensions[col].width = w

    # ── Header section (rows 1–7) ─────────────────────────────────────────────

    # Row 1: Event banner
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = (
        "MINI PROJECT EXHIBITION 2026  |  Dynamic System Modeling (702MH0C023)  |  "
        "B.Tech Mechatronics Engineering, Semester IV, AY 2025-26  |  NMIMS MPSTME"
    )
    ap(c,
       font=Font(bold=True, size=10, color="FFFFFF"),
       fill=hfill(H),
       align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 20

    # Row 2: Evaluator name & role
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"EVALUATOR:   {ev['name']}          ROLE:   {ev['role']}"
    ap(c,
       font=Font(bold=True, size=13, color="FFFFFF"),
       fill=hfill(H),
       align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[2].height = 30

    # Row 3: Session / date / venue
    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value = (
        f"Session:  {ev['session']}     "
        "Date:  25 April 2026 (Saturday)     "
        "Venue:  AR/VR Lab, MPSTME, NMIMS"
    )
    ap(c,
       font=Font(bold=True, size=9, color="FFFFFF"),
       fill=hfill(H),
       align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[3].height = 24

    # Row 4: Assessment focus
    ws.merge_cells("A4:H4")
    c = ws["A4"]
    c.value = "ASSESSMENT FOCUS:   " + ev["focus"]
    ap(c,
       font=Font(italic=True, size=9, color="333333"),
       fill=hfill(A),
       align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[4].height = 20

    # Row 5: SF-PBL context note
    ws.merge_cells("A5:H5")
    c = ws["A5"]
    c.value = ev["sfpbl"]
    ap(c,
       font=Font(bold=True, italic=True, size=9, color=H),
       fill=hfill("FFFFFF"),
       align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[5].height = 18

    # Row 6: Criteria key
    ws.merge_cells("A6:H6")
    key = "  |  ".join(
        [cr[0] + ": " + cr[1].split("[")[0].strip() for cr in ev["criteria"]]
    )
    c = ws["A6"]
    c.value = "CRITERIA KEY:   " + key
    ap(c,
       font=Font(bold=True, size=8, color="111111"),
       fill=hfill(A),
       align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[6].height = 22

    # Row 7: Column headers
    hdrs = [
        "Roll No.", "Student Name",
        f"C1\n/{max_c}", f"C2\n/{max_c}", f"C3\n/{max_c}", f"C4\n/{max_c}",
        f"Total\n/{tmax}",
        "Comments / Feedback",
    ]
    for ci, h in enumerate(hdrs, start=1):
        c = ws.cell(row=7, column=ci, value=h)
        ap(c,
           font=Font(bold=True, size=9, color="FFFFFF"),
           fill=hfill(H),
           align=Alignment(horizontal="center", vertical="center", wrap_text=True),
           border=thin())
    ws.row_dimensions[7].height = 34

    # ── Data rows (row 8 onwards) ─────────────────────────────────────────────
    r = DATA_START
    for g_idx, grp in enumerate(GROUPS):

        # Group header row
        ws.merge_cells(f"A{r}:H{r}")
        c = ws[f"A{r}"]
        c.value = f"  Group {grp['num']}   |   {grp['title']}"
        ap(c,
           font=Font(bold=True, size=9, color="FFFFFF"),
           fill=hfill(H),
           align=Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[r].height = 15
        r += 1

        # Student rows
        rfill = hfill("F7F9FC") if g_idx % 2 == 0 else hfill("FFFFFF")
        n = len(grp["students"])
        for s_idx, (roll, name) in enumerate(grp["students"]):
            last = (s_idx == n - 1)
            bdr  = sep() if last else thin()

            # Roll No (col 1)
            c = ws.cell(row=r, column=1, value=roll)
            ap(c, font=Font(size=9, bold=True), fill=rfill,
               align=Alignment(horizontal="center", vertical="center"), border=bdr)

            # Student Name (col 2)
            c = ws.cell(row=r, column=2, value=name)
            ap(c, font=Font(size=10, bold=True), fill=rfill,
               align=Alignment(horizontal="left", vertical="center"), border=bdr)

            # Score columns C1–C4 (cols 3–6)
            for ci in range(3, 7):
                c = ws.cell(row=r, column=ci, value="")
                ap(c, font=Font(size=13, bold=True), fill=S_FILL,
                   align=Alignment(horizontal="center", vertical="center"), border=bdr)

            # Total column G (col 7)
            total_val = f"=SUM(C{r}:F{r})" if use_formula else ""
            c = ws.cell(row=r, column=7, value=total_val)
            ap(c, font=Font(size=12, bold=True, color="155724"), fill=T_FILL,
               align=Alignment(horizontal="center", vertical="center"), border=bdr)

            # Comments column H (col 8)
            c = ws.cell(row=r, column=8, value="")
            ap(c, font=Font(size=8), fill=rfill,
               align=Alignment(horizontal="left", vertical="top", wrap_text=True), border=bdr)

            ws.row_dimensions[r].height = 22
            r += 1

    # ── Footer section ────────────────────────────────────────────────────────

    # Grading scale + GROUP/INDIVIDUAL guidance
    scale_row = r
    ws.merge_cells(f"A{scale_row}:H{scale_row}")
    c = ws[f"A{scale_row}"]
    c.value = _grading_scale_note(ev)
    ap(c,
       font=Font(italic=True, size=8, color="333333"),
       fill=hfill(A),
       align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[scale_row].height = 40

    # Criteria detail
    crit_row = scale_row + 1
    ws.merge_cells(f"A{crit_row}:H{crit_row}")
    detail = "  |  ".join(
        [f"{cr[0]} — {cr[1]}:  {cr[2]}" for cr in ev["criteria"]]
    )
    c = ws[f"A{crit_row}"]
    c.value = "CRITERIA DETAIL:   " + detail
    ap(c,
       font=Font(size=8, color="333333"),
       fill=hfill(A),
       align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[crit_row].height = 50

    # Signature line
    sig_row = crit_row + 1
    ws.merge_cells(f"A{sig_row}:B{sig_row}")
    c = ws[f"A{sig_row}"]
    c.value = f"Evaluator Signature:   {ev['name']}"
    ap(c, font=Font(bold=True, size=11), fill=hfill("FFFFFF"),
       align=Alignment(horizontal="left", vertical="center"), border=thin())

    ws.merge_cells(f"C{sig_row}:D{sig_row}")
    c = ws[f"C{sig_row}"]
    c.value = "Date:   25 / 04 / 2026"
    ap(c, font=Font(bold=True, size=11), fill=hfill("FFFFFF"),
       align=Alignment(horizontal="center", vertical="center"), border=thin())

    ws.merge_cells(f"E{sig_row}:H{sig_row}")
    c = ws[f"E{sig_row}"]
    c.value = "Coordinated by:   Prof. Sunny Nanade  (Course Instructor, NMIMS)"
    ap(c, font=Font(bold=True, size=11), fill=hfill("FFFFFF"),
       align=Alignment(horizontal="center", vertical="center"), border=thin())
    ws.row_dimensions[sig_row].height = 28

    # Print settings
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area              = f"A1:H{sig_row}"
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.3
    )
    ws.print_title_rows      = "7:7"
    ws.oddHeader.center.text = (
        f"Mini Project Exhibition 2026  |  {ev['name']}  |  25 April 2026  |  /{ev['total_max']} marks"
    )
    ws.oddFooter.right.text  = "Page &P of &N"
    ws.sheet_properties.tabColor = ev["tab_color"]


# =============================================================================
# MARKS SUMMARY SHEET  (11 cols A–K)
#
#   A  Grp #          (5)
#   B  Roll No.       (10)
#   C  Student Name   (22)
#   D  Project Title  (32)
#   E  Kamlesh  /16   (11)
#   F  Hasti    /16   (11)
#   G  Nitu     /16   (11)
#   H  Kane     /20   (11)
#   I  Parminder/16   (11)
#   J  Dash     /16   (11)
#   K  FINAL    /100  (13)
# =============================================================================

def build_summary_sheet(ws, ev_sheet_names):
    HDR     = "2C3E50"
    ACC     = "EBF5FB"
    J_FILLS = ["D6E4F0", "EFE0F5", "D5F0E2", "FFF3CD", "FAE5DC", "E8EAF6"]

    for col, w in {"A": 5, "B": 10, "C": 22, "D": 32,
                   "E": 11, "F": 11, "G": 11, "H": 11,
                   "I": 11, "J": 11, "K": 13}.items():
        ws.column_dimensions[col].width = w

    # Row 1: Banner
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = (
        "MINI PROJECT EXHIBITION 2026  --  MARKS SUMMARY  |  Final Score out of 100"
    )
    ap(c,
       font=Font(bold=True, size=13, color="FFFFFF"),
       fill=hfill(HDR),
       align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 26

    # Row 2: Marks breakdown
    ws.merge_cells("A2:K2")
    c = ws["A2"]
    breakdown = "  +  ".join(
        [f"{ev['name'].split()[-1]} /{ev['total_max']}" for ev in EVALUATORS]
    )
    c.value = (
        f"6 Judges = 100 Total  |  {breakdown}  =  FINAL /100\n"
        "Columns E–J are pulled automatically from evaluator sheets. "
        "Enter marks ONLY on the individual evaluator tabs (1–6)."
    )
    ap(c,
       font=Font(bold=True, italic=True, size=9, color="333333"),
       fill=hfill(ACC),
       align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[2].height = 34

    # Row 3: Column headers
    col_hdrs = (
        ["Grp\n#", "Roll\nNo.", "Student Name", "Project Title"]
        + [ev["col_hdr"] for ev in EVALUATORS]
        + ["FINAL\nMARKS\n/100"]
    )
    for ci, h in enumerate(col_hdrs, start=1):
        c = ws.cell(row=3, column=ci, value=h)
        ap(c,
           font=Font(bold=True, size=9, color="FFFFFF"),
           fill=hfill(HDR),
           align=Alignment(horizontal="center", vertical="center", wrap_text=True),
           border=thin())
    ws.row_dimensions[3].height = 46

    # Student data rows (rows 4 onward)
    total = len(STUDENT_ROW_MAP)
    for idx, s in enumerate(STUDENT_ROW_MAP):
        dr     = 4 + idx
        ev_row = s["row"]   # same row number in every evaluator sheet
        is_last_grp = (
            idx == total - 1 or
            STUDENT_ROW_MAP[idx + 1]["grp_num"] != s["grp_num"]
        )
        bdr      = sep() if is_last_grp else thin()
        row_fill = hfill("F7F9FC") if s["grp_num"] % 2 == 0 else hfill("FFFFFF")

        # Static columns A–D
        for ci, val in enumerate(
            [s["grp_num"], s["roll"], s["name"],
             s["title"][:50] + ("..." if len(s["title"]) > 50 else "")],
            start=1
        ):
            c = ws.cell(row=dr, column=ci, value=val)
            ap(c,
               font=Font(size=9, bold=(ci <= 2)),
               fill=row_fill,
               align=Alignment(
                   horizontal="center" if ci <= 2 else "left",
                   vertical="center", wrap_text=(ci == 4)),
               border=bdr)

        # Judge totals E–J: cross-sheet formula to col G of each evaluator sheet
        for ji, sheet in enumerate(ev_sheet_names):
            c = ws.cell(row=dr, column=5 + ji,
                        value=f"='{sheet}'!G{ev_row}")
            ap(c,
               font=Font(size=11, bold=True),
               fill=hfill(J_FILLS[ji]),
               align=Alignment(horizontal="center", vertical="center"),
               border=bdr)

        # FINAL col K (column 11)
        c = ws.cell(row=dr, column=11,
                    value=f"=E{dr}+F{dr}+G{dr}+H{dr}+I{dr}+J{dr}")
        ap(c,
           font=Font(size=13, bold=True, color="FFFFFF"),
           fill=hfill(HDR),
           align=Alignment(horizontal="center", vertical="center"),
           border=bdr)

        ws.row_dimensions[dr].height = 20

    # Class average row
    avg_row = 4 + total
    ws.merge_cells(f"A{avg_row}:D{avg_row}")
    c = ws[f"A{avg_row}"]
    c.value = "CLASS AVERAGE"
    ap(c,
       font=Font(bold=True, size=10, color="FFFFFF"),
       fill=hfill(HDR),
       align=Alignment(horizontal="center", vertical="center"),
       border=thin())
    for col in range(5, 12):
        cl = get_column_letter(col)
        c = ws.cell(row=avg_row, column=col,
                    value=f"=ROUND(AVERAGE({cl}4:{cl}{avg_row-1}),2)")
        ap(c,
           font=Font(bold=True, size=11, color="FFFFFF"),
           fill=hfill(HDR),
           align=Alignment(horizontal="center", vertical="center"),
           border=thin())
    ws.row_dimensions[avg_row].height = 24

    # Note row
    note_row = avg_row + 1
    ws.merge_cells(f"A{note_row}:K{note_row}")
    c = ws[f"A{note_row}"]
    c.value = (
        "NOTE:  C1-C3 are typically GROUP criteria (same score for all students in a group).  "
        "C4 is typically INDIVIDUAL (each student scored separately).  "
        "Parminder's marks combine morning demo + afternoon formal presentation.  "
        "Do NOT edit columns E-J here — enter all marks on the individual evaluator tabs.  MAX = 100."
    )
    ap(c,
       font=Font(italic=True, size=8, color="444444"),
       fill=hfill(ACC),
       align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[note_row].height = 28

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows        = "3:3"
    ws.sheet_properties.tabColor = "2C3E50"


# =============================================================================
# MAIN
# =============================================================================
def main():
    tab_labels = [
        "1_Kamlesh_Panchal",
        "2_Hasti_Chandarana",
        "3_Nitu_Gupta",
        "4_Mahendra_Kane",
        "5_Parminder_Jandoo",
        "6_Debasis_Dash",
    ]

    # ── Digital version: SUM formulas + summary sheet ─────────────────────────
    print("Building digital version (with formulas)...")
    wb1 = Workbook()
    wb1.remove(wb1.active)
    sheet_names = []
    for i, ev in enumerate(EVALUATORS):
        ws = wb1.create_sheet(title=tab_labels[i])
        build_evaluator_sheet(ws, ev, use_formula=True)
        sheet_names.append(tab_labels[i])
        print(f"  + {tab_labels[i]}  (/{ev['total_max']} marks)")

    ws_sum = wb1.create_sheet(title="7_Marks_Summary")
    build_summary_sheet(ws_sum, sheet_names)
    print("  + 7_Marks_Summary")

    out1 = os.path.join(OUT_DIR, "Evaluation_Rubric_Sheets.xlsx")
    wb1.save(out1)
    print(f"\n  [DIGITAL] Saved: {out1}")

    # ── Print version: blank total cells, 6 evaluator sheets only ─────────────
    print("\nBuilding print version (blank totals, no summary)...")
    wb2 = Workbook()
    wb2.remove(wb2.active)
    for i, ev in enumerate(EVALUATORS):
        ws = wb2.create_sheet(title=tab_labels[i])
        build_evaluator_sheet(ws, ev, use_formula=False)
        print(f"  + {tab_labels[i]}")

    out2 = os.path.join(OUT_DIR, "Evaluation_Rubric_Sheets_PRINT.xlsx")
    wb2.save(out2)
    print(f"\n  [PRINT]   Saved: {out2}")

    # ── Summary ───────────────────────────────────────────────────────────────
    total_marks = sum(ev["total_max"] for ev in EVALUATORS)
    print(f"\n  MARKS STRUCTURE:  {len(EVALUATORS)} Judges = {total_marks} Total")
    print(f"  {'─'*54}")
    for ev in EVALUATORS:
        print(f"  {ev['name']:<35}  /{ev['total_max']}")
    print(f"  {'─'*54}")
    print(f"  {'FINAL MARKS':<35}  /{total_marks}")
    print(f"\n  {len(GROUPS)} groups  |  {len(STUDENT_ROW_MAP)} individual student rows")
    print(f"  C1–C3: GROUP criteria  |  C4: INDIVIDUAL criterion  (see each sheet for exact labels)")


if __name__ == "__main__":
    main()
